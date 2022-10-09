from enum import IntEnum, auto
import pyperclip
from selenium import webdriver
from chromedriver_autoinstaller import install
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import openpyxl
import os
import datetime
from openpyxl.drawing.image import Image
import urllib.request as req

class Role(IntEnum):
    전체직업 = 0
    버서커 = auto()
    디스트로이어 = auto()
    워로드 = auto()
    홀리나이트 = auto()
    아르카나 = auto()
    서머너 = auto()
    바드 = auto()
    소서리스 = auto()
    배틀마스터 = auto()
    인파이터 = auto()
    기공사 = auto()
    창술사 = auto()
    스트라이커 = auto()
    블레이드 = auto()
    데모닉 = auto()
    리퍼 = auto()
    호크아이 = auto()
    데빌헌터 = auto()
    블래스터 = auto()
    스카우터 = auto()
    건슬링어 = auto()
    도화가 = auto()
    기상술사 = auto()

avatar_list = [
    'S',
    '3주년 미래',
    '3주년 미래 얼굴2',
    '3주년 미래 머리',
    '3주년 미래 하의',
    '3주년 미래 상의',
    '3주년 미래 데모닉웨폰',
    'S',
    '3주년 추억',
    '3주년 추억 얼굴2',
    '3주년 추억 머리',
    '3주년 추억 하의',
    '3주년 추억 상의',
    '3주년 추억 데모닉웨폰'
    'S',
    '테돈바드',
    '테돈바드 비치웨어 머리',
    '테돈바드 비치웨어 상의',
    '테돈바드 비치웨어 하의'
]





def input_id_pw(browser, css, id_or_pw):
    pyperclip.copy(id_or_pw)  # id 복사
    blank = browser.find_element_by_css_selector(css)
    blank.click()
    ActionChains(browser).key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()

    time.sleep(1)  # 완전 범죄

# -----------------------------------------------------------------------------------------------------
# 엑셀 시트
# 이미지 저장할 폴더 생성
if not os.path.exists("./로스트아크_아바타_이미지"):
    os.mkdir("./로스트아크_아바타_이미지")

file_name = "./로스트아크_아바타_가격.xlsx"
# 엑셀 파일이 존재 하지 않으면 엑셀 파일 생성
if not os.path.exists(file_name):
    openpyxl.Workbook().save(file_name)
# 엑셀 파일 불러 오기
book = openpyxl.load_workbook(file_name)
# 쓸데 없는 시트 지우기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
# 새로운 시트 만들고, 시트 이름 짓기
sheet = book.create_sheet()
sheet.title = datetime.datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
sheet.column_dimensions["A"].width = 16
sheet.column_dimensions["B"].width = 12
sheet.column_dimensions["C"].width = 12
sheet.column_dimensions["D"].width = 12
# -----------------------------------------------------------------------------------------------------
# 크롤링
browser = webdriver.Chrome(install())
browser.implicitly_wait(3)
browser.get("https://member.onstove.com/auth/login?inflow_path=lost_ark&game_no=45&redirect_url=https%3a%2f%2flostark.game.onstove.com%2fMarket")
time.sleep(1)

input_id_pw(browser, "input#user_id", "osj5137@naver.com")
input_id_pw(browser, "input#user_pwd", "T+!,,y=]y,**mGwRn77*v!B?")

login_button = browser.find_element_by_css_selector("div.row.grid.el-actions > button.el-btn.btn-primary.size-56.round.width-full")
login_button.click()
time.sleep(3)

select_class_blank = browser.find_element_by_css_selector("div.class div.lui-select__title")
select_class_blank.click()

select_classes = browser.find_elements_by_css_selector("div.class div.lui-select__option label")
time.sleep(0.1)
select_classes[Role.데모닉].click()
set_name = False
excel_cell_num = 1
img_num = 1
for avatar_num in range(len(avatar_list)):
    print(avatar_list[avatar_num])
    if avatar_list[avatar_num] == 'S':
        set_name = True
        excel_cell_num += 1
        continue
    elif set_name == True:
        sheet.cell(row=excel_cell_num + 1, column= 1).value = avatar_list[avatar_num]
        sheet.cell(row=excel_cell_num + 1, column= 2).value = "최저가(골드)"
        sheet.cell(row=excel_cell_num + 1, column= 3).value = "평균가(골드)"
        sheet.cell(row=excel_cell_num + 1, column= 4).value = "거래가능회수"
        excel_cell_num += 1
        set_name = False
        continue
    search_item = browser.find_element_by_css_selector("input#txtItemName")
    search_item.clear()
    search_item.send_keys(avatar_list[avatar_num])
    ActionChains(browser).key_down(Keys.ENTER).perform()
    time.sleep(1)
    try:
        browser.find_element_by_css_selector("a.pagination__last").click()
    except:
        print("lastPage!")
    time.sleep(2)
    current_page = browser.find_element_by_css_selector("em.pagination__number.pagination__number--active")
    page_num = int(current_page.text)
    while True:
        item_list = reversed(browser.find_elements_by_css_selector("tbody#tbodyItemList tr"))
        for item in item_list:
            time.sleep(0.1)
            img = item.find_element_by_css_selector("img")
            img_url = img.get_attribute('src')
            time.sleep(0.1)
            current_lowest_price_list = item.find_elements_by_css_selector("div.price>em") # 최저가
            current_lowest_price = current_lowest_price_list[2].text
            print(f"최저가 : {current_lowest_price}")
            item.find_element_by_css_selector("button.button.button--deal-price").click()
            time.sleep(0.4)
            average_price_list = browser.find_elements_by_css_selector("div.info-box tbody div.price>em") # 평균가
            average_price = average_price_list[1].text
            print(f"평균가 : {average_price}")
            time.sleep(0.1)
            trade_cnt = item.find_element_by_css_selector("span.count em").text # 거래가능 횟수
            print(f"거래횟수 : {trade_cnt}")
            close_btn = browser.find_element_by_xpath('//*[@id="modal-deal-price"]/div/div/button')
            browser.execute_script("arguments[0].click()", close_btn)

            img_file_name = f"./로스트아크_아바타_이미지/{img_num}.png"
            req.urlretrieve(img_url, img_file_name)  # 이미지 다운로드
            # 엑셀에 이미지 넣기
            img_for_excel = Image(img_file_name)
            sheet.add_image(img_for_excel, f"A{excel_cell_num + 1}")
            sheet.row_dimensions[excel_cell_num + 1].height = 100
            # 엑셀에 크롤링한 데이터 넣기
            print(excel_cell_num)
            sheet.cell(row=excel_cell_num + 1, column=2).value = current_lowest_price
            sheet.cell(row=excel_cell_num + 1, column=3).value = average_price
            sheet.cell(row=excel_cell_num + 1, column=4).value = trade_cnt
            excel_cell_num += 1

        img_num += 1
        page_num -= 1
        page_buttons = browser.find_elements_by_css_selector("a.pagination__number")
        for i in page_buttons:
            if int(i.text) == page_num:
                i.click()
                time.sleep(1)
                break
        else:
            if page_num == 0:
                break
            next_page_button = browser.find_element_by_xpath('//*[@id="marketList"]/div[2]/a[2]')
            next_page_button.click()
            time.sleep(1)

book.active = sheet
book.save(file_name)
browser.close()