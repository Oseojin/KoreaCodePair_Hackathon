import datetime
from bs4 import BeautifulSoup
import urllib.request as req
import openpyxl
import os
from openpyxl.drawing.image import Image

# 이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

file_name = "./멜론_크롤링.xlsx"
# 엑셀 파일이 존재하지 않으면 엑셀 파일 생성
if not os.path.exists(file_name):
    openpyxl.Workbook().save(file_name)
# 엑셀 파일 불러오기
book = openpyxl.load_workbook(file_name)
# 쓸데없는 시트 지우기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
# 새로운 시트 만들고, 시트 이름 짓기
sheet = book.create_sheet()
sheet.title = datetime.datetime.today().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
# 열 너비 조정
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 36
sheet.column_dimensions["C"].width = 30
sheet.column_dimensions["D"].width = 45

header = req.Request("https://www.melon.com/chart/", headers={"User-Agent" : "Mozilla/5.0"})
code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")

title = soup.select("div.wrap_song_info > div.ellipsis.rank01 a")
name = soup.select("div.ellipsis.rank02 > span")
album = soup.select("div.ellipsis.rank03 > a")
img = soup.select("a.image_typeAll > img")

for i in range(len(title)):
    data_title = title[i].text
    data_name = name[i].text
    data_album = album[i].text
    img_url = img[i].attrs["src"]
    img_file_name = f"./멜론이미지/{i + 1}.png"

    print(f"{i+1}위 : {data_title} - {data_name} / {data_album}")
    print(img_url) # 이미지 URL 주소 추출
    
    req.urlretrieve(img_url, img_file_name) # 이미지 다운로드
    # 엑셀에 이미지 넣기
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel, f"A{i+1}")
    sheet.row_dimensions[i+1].height = 92

    # 엑셀에 크롤링한 데이터 넣기
    sheet.cell(row=i+1, column=2).value = data_title
    sheet.cell(row=i+1, column=3).value = data_name
    sheet.cell(row=i+1, column=4).value = data_album

book.active = sheet
book.save(file_name)